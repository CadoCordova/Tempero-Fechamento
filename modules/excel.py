from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def formatar_tabela_excel(ws, df, start_row=1):
    """
    Aplica estilo básico à tabela:
    - Cabeçalho em negrito, fundo cinza, centralizado
    - Largura das colunas ajustada ao conteúdo
    - Colunas de valor com formato de moeda (R$)
    """
    n_rows = len(df)
    n_cols = len(df.columns)

    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = ws[f"A{start_row + 1}"]

    for col_idx in range(1, n_cols + 1):
        max_len = 0
        for row_idx in range(start_row, start_row + 1 + n_rows):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    col_names_lower = [str(c).lower() for c in df.columns]
    currency_prefixes = ("entradas", "saídas", "saidas", "resultado", "saldo", "valor")
    for col_idx, col_name in enumerate(col_names_lower, start=1):
        if any(col_name.startswith(p) for p in currency_prefixes):
            for row_idx in range(start_row + 1, start_row + 1 + n_rows):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"R$" #,##0.00'
